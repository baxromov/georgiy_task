from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import ssl
from datetime import datetime
import email
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook
from apscheduler.schedulers.blocking import BlockingScheduler
import email.mime.application
sched = BlockingScheduler()


@sched.scheduled_job('interval', minutes=1)
def send_email():
    msg = MIMEMultipart("alternative")
    fromaddr = "bukanov1234@mail.ru"
    mypass = "cRYfj13YTp65wmluZxJU"
    toaddr = "bukanov1234@mail.ru"
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "Отправитель: Telegram bot"

    now = datetime.now()
    response_date = now.strftime("%d.%m.%Y")

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
    </head>
    <body>        
    <h1>Отчёт за: {response_date} </h1>      
    </body>
    </html>
    """
    text = bs(html, "html.parser").text
    msg.attach(MIMEText(text, 'plain'))
    msg.attach(MIMEText(html, 'html', 'utf-8'))

    filename = 'bot/data/example.xlsx'
    fp = open(filename, 'rb')
    att = email.mime.application.MIMEApplication(fp.read(), _subtype="xlsx")
    fp.close()
    att.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(att)

    server = smtplib.SMTP_SSL('smtp.mail.ru:465')
    context = ssl.SSLContext(ssl.PROTOCOL_TLS)
    server.login(msg['From'], mypass)
    text = msg.as_string()
    server.sendmail(msg['From'], msg['To'], text)
    server.quit()

    print("Successfully")
    clear_sheet()


def clear_sheet():
    fn = 'bot/data/example.xlsx'
    wb = load_workbook(fn)
    ws = wb['Лист1']
    nb_row = ws.max_row
    ws.delete_rows(2, nb_row)
    wb.save('example.xlsx')
